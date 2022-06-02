from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('tim.xlsx')
ws = wb.active

ws.move_range("A1:D8", rows=2, cols=2)

wb.save('tim.xlsx')
