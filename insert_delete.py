from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('tim.xlsx')
ws = wb.active

# ws.insert_rows(7)
# ws.delete_rows(7)

# ws.insert_cols(2)
# ws.insert_cols(2)
ws.delete_cols(2)

wb.save('tim.xlsx')
